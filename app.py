import streamlit as st
import pandas as pd
from preprocess import build_model_input
from optimiser import solve_week
import helper_maxflow
import tempfile
import time
from collections import defaultdict
import io
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Custom ARISE Theme CSS ---
st.markdown("""
    <style>
    .stApp {
        background-color: #fff !important;
        color: #111 !important;
    }
    h1, h2, h3, h4, h5, h6 {
        color: #111 !important;
        margin-top: 0.5em;
        margin-bottom: 0.5em;
    }
    .stMetric, .stAlert, .stDataFrame {
        background-color: #fff !important;
        color: #111 !important;
        border-radius: 14px;
        padding: 1.5em 1.5em 1.5em 1.5em !important;
        margin-bottom: 1.5em !important;
        border: 2px solid #388e3c22;
        box-shadow: 0 2px 12px 0 #1976d233;
    }
    .stMetric {
        border-left: 6px solid #388e3c;
        margin-right: 1em;
    }
    .stDataFrame {
        border-left: 6px solid #1976d2;
    }
    .stButton>button {
        background-color: #388e3c !important;
        color: #fff !important;
        border-radius: 8px;
        border: none;
        padding: 0.75em 2em;
        font-size: 1.1em;
        margin-top: 0.5em;
        margin-bottom: 0.5em;
    }
    .stButton>button:hover {
        background-color: #11f8f4 !important; /* Teal on hover */
        color: #11f8f4 !important;
    }
    .stAlert {
        background-color: transparent !important;
        color: #11f8f4 !important;
        border: none !important;
        box-shadow: none !important;
        padding: 1.2em 1.2em 1.2em 1.2em !important;
    }
    .stFileUploader, .stSelectbox {
        background-color: #fff !important;
        color: #089258 !important;
        border-radius: 10px;
        padding: 1em 1em 1em 1em !important;
    }
    /* Remove top padding/margin from the main container */
    .block-container {
        padding-top: 0.5rem !important;
        margin-top: 0 !important;
    }
    .main {
        padding-top: 0 !important;
        margin-top: 0 !important;
    }
    /* Reduce margin below the subtitle */
    .block-container > div:nth-child(4) {
        background: transparent !important;
        box-shadow: none !important;
        border: none !important;
        padding: 0 !important;
        margin-bottom: 0.2em !important;
    }
    .custom-breakdown-container {
        background: #fff;
        border-left: 6px solid #1976d2;  /* Blue strip, use #388e3c for green */
        border-radius: 14px;
        box-shadow: 0 2px 12px 0 #1976d233;
        padding: 1.5em 1.5em 1.5em 1.5em;
        margin-bottom: 1.5em;
        margin-top: 0.5em;
    }
    </style>
""", unsafe_allow_html=True)

# --- ARISE Logo (update path as needed) ---
st.image("arise_logo.png", width=180)

# --- Ship/Boat Emoji Animation ---
ship_emojis = ["🎄"]

st.set_page_config(page_title="Truck Forest Allocation Optimizer", layout="wide")
st.title("🪵Truck Forest Allocation Optimizer")
st.markdown("<span style='font-size:1.2em; color:#90caf9; font-weight:500;'>Optimize truck allocations for maximum CBM collection from forests to NKOK, Libreville</span>", unsafe_allow_html=True)

# --- Inputs ---
col1, col2, col3 = st.columns(3)
with col1:
    forests_file = st.file_uploader("Upload Forests CSV", type="csv")
with col2:
    trucks_file = st.file_uploader("Upload Trucks CSV", type="csv")
with col3:
    season = st.selectbox("Season", ["dry", "rain"])

# --- User input for cost per CBM ---
# (No longer needed, cost is not used)

if forests_file and trucks_file:
    # --- Animated Title While Loading ---
    title_placeholder = st.empty()
    start_time = time.time()
    for i in range(12):  # ~1.2s animation (adjust as needed)
        title_placeholder.title(f"{ship_emojis[i % len(ship_emojis)]} Truck Forest Allocation Optimizer")
        time.sleep(0.1)
    # Save uploaded files to temp files for pipeline compatibility
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as f1, \
         tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as f2:
        f1.write(forests_file.getbuffer())
        f2.write(trucks_file.getbuffer())
        forests_path = f1.name
        trucks_path = f2.name

    # Build model input
    df = build_model_input(forests_csv=forests_path, trucks_csv=trucks_path, season=season)

    # Run optimizer (maximize profit)
    plan = solve_week(df, maximize_profit=True)
    elapsed = time.time() - start_time
    # Show final emoji after loading
    title_placeholder.title(f"{ship_emojis[int(time.time()) % len(ship_emojis)]} Truck Forest Allocation Optimizer")

    # Ensure 'hours_used' exists in plan
    if 'hours_used' not in plan.columns:
        plan['hours_used'] = plan.apply(lambda row: row['trips_planned'] * df.loc[(row['truck_id'], row['forest_id']), 'trip_hours'], axis=1)

    # Ensure 'profit' column exists in plan
    if 'profit' not in plan.columns:
        plan['profit'] = plan.apply(lambda row: row['trips_planned'] * df.loc[(row['truck_id'], row['forest_id']), 'cbm_per_truck'] * df.loc[(row['truck_id'], row['forest_id']), 'profit_per_cbm_euros'], axis=1)

    # --- After plan is created, compute per-forest volumes and remaining_by_forest ---
    forest_volumes = df.reset_index().drop_duplicates("forest_id").set_index("forest_id")["weekly_stockpile_cbm"]
    plan["depleted_cbm"] = plan["trips_planned"] * plan["cbm_per_truck"]
    depleted_by_forest = plan.groupby("forest_id")["depleted_cbm"].sum()
    remaining_by_forest = {forest: forest_volumes[forest] - depleted_by_forest.get(forest, 0) for forest in forest_volumes.index}

    # Aggregate allocations
    allocations = plan.groupby("forest_id").agg(
        CBM=("depleted_cbm", "sum"),
        Trips=("trips_planned", "sum"),
        Hours=("hours_used", "sum"),
        Profit=("profit", "sum")
    ).reset_index()
    allocations = allocations.rename(columns={"forest_id": "Forest"})
    allocations["Remaining"] = allocations["Forest"].map(remaining_by_forest)
    allocations["Efficiency (CBM/hr)"] = allocations.apply(
        lambda row: row["CBM"] / row["Hours"] if row["Hours"] > 0 else 0, axis=1
    )
    # After allocations groupby, ensure all forests are shown in the table
    all_forests = pd.DataFrame({"Forest": df.reset_index()["forest_id"].unique()})
    allocations = allocations.merge(all_forests, on="Forest", how="right")
    # Fill NaN for numeric columns with 0
    for col in ["Volume collected", "Trips", "Hours used", "Profit", "Remaining", "Efficiency (CBM/hr)"]:
        if col in allocations.columns:
            allocations[col] = allocations[col].fillna(0)
    # Merge truck lists into allocations and fill NaN with ''
    truck_lists = plan.groupby("forest_id")["truck_id"].apply(lambda x: ', '.join(map(str, x.unique()))).reset_index()
    truck_lists = truck_lists.rename(columns={"forest_id": "Forest", "truck_id": "Trucks Assigned"})
    allocations = allocations.merge(truck_lists, on="Forest", how="left")
    if "Trucks Assigned" in allocations.columns:
        allocations["Trucks Assigned"] = allocations["Trucks Assigned"].fillna("")
    # Partially harvested forests
    partials = allocations[allocations["Remaining"] > 0]

    # --- Max-flow (full trips) ---
    trucks = pd.read_csv(trucks_path)
    assigned_trucks = set(plan["truck_id"].astype(str))
    unassigned = trucks[~trucks["truck_id"].astype(str).isin(assigned_trucks)]
    idle_df = unassigned.copy()
    idle_df["available_hours"] = idle_df["drive_hours"] - idle_df.get("maintenance_hours", 0)
    # Include all forests with volume > 0 and profit_per_cbm_euros > 0
    forests_with_leftover = [
        f for f in forest_volumes.index
        if forest_volumes[f] > 0 and (df.reset_index().set_index('forest_id').loc[f, 'profit_per_cbm_euros'] > 0).any()
    ]
    forests_df = (
        df.reset_index()
          .drop_duplicates("forest_id")
          .set_index("forest_id")
          .loc[forests_with_leftover]
          .reset_index()[["forest_id", "trip_hours", "cbm_per_truck", "profit_per_cbm_euros"]]
    )
    forests_df = forests_df.rename(columns={"trip_hours": "turnaround_time"})
    forests_df["volume_left"] = forests_df["forest_id"].map(remaining_by_forest)
    forests_df["profit_per_trip"] = forests_df["cbm_per_truck"] * forests_df["profit_per_cbm_euros"]
    # Pass profit_per_trip to helper_maxflow
    extra_assignments = helper_maxflow.top_up_with_flow(idle_df, forests_df)
    # --- Safe max-flow assignment handling (reflecting scratch.py logic) ---
    rows = []
    for a in extra_assignments:
        key = (int(a["truck_id"]), a["forest_id"])
        if key in df.index:
            cbm_per_truck = df.loc[key, "cbm_per_truck"]
            rows.append({
                "truck_id": a["truck_id"],
                "forest_id": a["forest_id"],
                "trips_planned": a["trips"],
                "cbm_per_truck": cbm_per_truck
            })
        # else: could log or st.write(f"Skipping invalid assignment: {key}")
    new_plan = pd.DataFrame(rows)
    if not new_plan.empty:
        st.subheader("🟢 Extra Assignments from Max-Flow (Full Trips)")
        st.dataframe(new_plan, use_container_width=True)
        st.write(f"Extra CBM from max-flow full trips: {new_plan['trips_planned'].mul(new_plan['cbm_per_truck']).sum():,.0f} m³")
    else:
        st.info("No valid assignments could be made in the second pass (max-flow, full trips).")

    half_assignments = helper_maxflow.half_trip_maxflow(idle_df, forests_df)
    half_cbm = sum(a["cbm_collected"] for a in half_assignments) if half_assignments else 0

    # --- Compute summary statistics for the summary dictionary ---
    total_cbm = (plan['trips_planned'] * plan['cbm_per_truck']).sum()
    total_trips = plan['trips_planned'].sum()
    trucks_used = plan['truck_id'].nunique()
    trucks_total = pd.read_csv(trucks_path)['truck_id'].nunique()
    trucks_unused = trucks_total - trucks_used

    # --- Output to Streamlit ---
    total_profit = allocations['Profit'].sum()
    summary = {
        "Total CBM": total_cbm,
        "Total Trips": total_trips,
        "Trucks Used": trucks_used,
        "Trucks Unused": trucks_unused,
        "Total Profit Euros": total_profit,
    }

    # --- Output: Summary Statistics ---
    st.subheader("📊 Summary Statistics")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total CBM", f"{summary['Total CBM']:,}")
    col2.metric("Total Trips", f"{summary['Total Trips']:,}")
    col3.metric("Trucks Used", f"{summary['Trucks Used']:,}")
    col4.metric("Trucks Unused", f"{summary['Trucks Unused']:,}")
    # Only show profit in a new row
    col5 = st.columns(1)[0]
    col5.metric("Total Profit (Euros)", f"€ {summary['Total Profit Euros']:,.0f}")

    # --- Output: Unused Trucks ---
    if not unassigned.empty:
        st.subheader("🚚 Unused Trucks")
        st.dataframe(unassigned[["truck_id", "type"]], use_container_width=True)
        unused_csv = unassigned[["truck_id", "type"]].to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download Unused Trucks as CSV",
            data=unused_csv,
            file_name="unused_trucks.csv",
            mime="text/csv",
        )

    # --- Output: Forest Allocations ---
    st.subheader("🌲 Forest Allocations")
    st.dataframe(allocations, use_container_width=True)

    # --- Add truck trip breakdown and total trips per forest ---
    truck_trip_breakdown = plan.groupby('forest_id').apply(
        lambda df_: ', '.join([f"Truck {int(row['truck_id'])}: {int(row['trips_planned'])}" for _, row in df_.iterrows()])
    ).reset_index().rename(columns={0: 'Truck Trip Breakdown', 'forest_id': 'Forest'})
    total_trips_per_forest = plan.groupby('forest_id')['trips_planned'].sum().reset_index().rename(columns={'forest_id': 'Forest', 'trips_planned': 'Total Trips'})
    breakdown_table = truck_trip_breakdown.merge(total_trips_per_forest, on='Forest', how='left')

    # --- Per-Forest Truck Trip Breakdown (HTML table with wrapping, no container) ---
    st.subheader("🚚 Per-Forest Truck Trip Breakdown")
    st.markdown(
        '<style>'
        '.custom-table td {'
        '    white-space: pre-wrap !important;'
        '    word-break: break-word !important;'
        '    max-width: 600px;'
        '    font-size: 1.1em;'
        '    padding: 10px 8px;'
        '}'
        '.custom-table th {'
        '    font-size: 1.1em;'
        '    padding: 10px 8px;'
        '}'
        '</style>'
        + breakdown_table.to_html(index=False, classes='custom-table', escape=False),
        unsafe_allow_html=True
    )
    # --- Download button for breakdown table as CSV ---
    breakdown_csv = breakdown_table.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="Download Truck Trip Breakdown as CSV",
        data=breakdown_csv,
        file_name="truck_trip_breakdown.csv",
        mime="text/csv",
    )

    # --- Function to generate daily forest-truck plan DataFrame from plan DataFrame ---
    def generate_daily_forest_plan(plan_df):
        trips_by_truck = defaultdict(list)
        for _, row in plan_df.iterrows():
            truck_id = str(row['truck_id']).strip()
            forest_id = str(row['forest_id']).strip()
            trips_planned = int(row['trips_planned'])
            cbm_per_truck = row['cbm_per_truck']
            profit = row['profit']
            trips_by_truck[truck_id].append({
                'forest_id': forest_id,
                'trips_planned': trips_planned,
                'cbm_per_truck': cbm_per_truck,
                'profit': profit
            })
        expanded_trips = []
        for truck_id, trips in trips_by_truck.items():
            total_trips = sum(trip['trips_planned'] for trip in trips)
            for trip in trips:
                for i in range(trip['trips_planned']):
                    expanded_trips.append({
                        'truck_id': truck_id,
                        'forest_id': trip['forest_id'],
                        'cbm_per_truck': trip['cbm_per_truck'],
                        'profit': trip['profit'],
                    })
        expanded_trips.sort(key=lambda x: int(x['truck_id']))
        trip_counters = defaultdict(int)
        for trip in expanded_trips:
            truck_id = trip['truck_id']
            trip_counters[truck_id] += 1
            trip['trip_number'] = trip_counters[truck_id]
            trip['day'] = trip_counters[truck_id]
        # Sort by day, then forest_id, then truck_id
        sorted_trips = sorted(expanded_trips, key=lambda x: (int(x['day']), str(x['forest_id']), int(x['truck_id'])))
        # Build DataFrame (without total_trips_for_truck)
        daily_forest_plan_df = pd.DataFrame(sorted_trips)[[
            'day', 'forest_id', 'truck_id', 'trip_number', 'cbm_per_truck', 'profit']]
        return daily_forest_plan_df

    # --- Generate daily forest-truck plan from plan DataFrame ---
    daily_forest_plan = generate_daily_forest_plan(plan)

    # --- Daily Forest-Truck Assignment Schedule (Grouped) ---
    st.subheader("🗓️ Daily Forest-Truck Assignment Schedule (Grouped)")
    try:
        grouped_lines = []
        for day, day_df in daily_forest_plan.groupby('day'):
            grouped_lines.append(f"<b>Day {int(day)}</b>")
            for forest, forest_df in day_df.groupby('forest_id'):
                trucks = [f"{row.truck_id} ({row.trip_number})" for row in forest_df.itertuples()]
                grouped_lines.append(f"&nbsp;&nbsp;<b>{forest}</b>: {', '.join(trucks)}<br>")
            grouped_lines.append("")
        st.markdown("<br>".join(grouped_lines), unsafe_allow_html=True)
    except Exception as e:
        st.warning(f"Could not generate grouped daily forest-truck schedule: {e}")

    # --- Add total trucks column to allocations ---
    if 'Trucks Assigned' in allocations.columns:
        allocations['Total Trucks'] = allocations['Trucks Assigned'].apply(lambda x: len([t for t in str(x).split(',') if t.strip()]))
    else:
        allocations['Total Trucks'] = 0

    # --- Generate grouped daily forest-truck summary for Excel ---
    def generate_grouped_daily_plan(daily_forest_plan):
        grouped = {}
        for day, day_df in daily_forest_plan.groupby('day'):
            day_df_sorted = day_df.sort_values(['forest_id', 'truck_id'])
            rows = []
            for forest, forest_df in day_df_sorted.groupby('forest_id'):
                truck_ids_list = [str(tid) for tid in forest_df['truck_id'].tolist()]
                truck_ids = ', '.join(truck_ids_list)
                total_trucks = len(truck_ids_list)
                rows.append({'Forest': forest, 'Truck IDs': truck_ids, 'Total Trucks': total_trucks})
            grouped[int(day)] = rows
        return grouped

    grouped_daily_plan = generate_grouped_daily_plan(daily_forest_plan)

    with io.BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Write Forest Allocations as usual, now with Total Trucks
            allocations.to_excel(writer, sheet_name='Forest Allocations', index=False)
            ws_alloc = writer.book['Forest Allocations']
            # Format Forest Allocations: text wrap, ample sizing, header color, borders
            header_fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
            header_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Header row
            for cell in ws_alloc[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            # Data rows: wrap, center, borders
            for row in ws_alloc.iter_rows(min_row=2, max_row=ws_alloc.max_row, min_col=1, max_col=ws_alloc.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.border = border
            # Zebra striping for data rows (alternate light blue and white)
            light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            for row_idx in range(2, ws_alloc.max_row + 1):
                if (row_idx - 1) % 2 == 1:
                    for cell in ws_alloc[row_idx]:
                        cell.fill = light_blue_fill
            # Column widths: set min width for key columns, auto-fit others, then add extra space
            col_map = {name: idx+1 for idx, name in enumerate(allocations.columns)}
            min_widths = {'Forest': 24, 'Profit': 18, 'Efficiency (CBM/hr)': 18, 'Trucks Assigned': 28, 'Total Trucks': 14}
            for col_name, min_w in min_widths.items():
                if col_name in col_map:
                    col_letter = ws_alloc.cell(row=1, column=col_map[col_name]).column_letter
                    ws_alloc.column_dimensions[col_letter].width = min_w
            for col in ws_alloc.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                if ws_alloc.column_dimensions[col_letter].width < max_length + 4:
                    ws_alloc.column_dimensions[col_letter].width = min(max_length + 4, 40)
                # Add extra width for breathing room
                ws_alloc.column_dimensions[col_letter].width += 7
            # Row heights: set all to 72 for maximum readability
            for row_idx in range(1, ws_alloc.max_row + 1):
                ws_alloc.row_dimensions[row_idx].height = 72

            # --- Write grouped daily plan as separate tables for each day ---
            ws = writer.book.create_sheet('Daily Forest-Truck Plan')
            start_row = 1
            for day in sorted(grouped_daily_plan.keys()):
                # Day label
                ws.cell(row=start_row, column=1, value=f"Day {day}")
                ws.cell(row=start_row, column=1).font = Font(bold=True, size=13)
                start_row += 1
                # Header
                ws.cell(row=start_row, column=1, value="Forest")
                ws.cell(row=start_row, column=2, value="Truck IDs")
                ws.cell(row=start_row, column=3, value="Total Trucks")
                for col in range(1, 4):
                    cell = ws.cell(row=start_row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                header_row = start_row
                start_row += 1
                # Table rows
                for row in grouped_daily_plan[day]:
                    ws.cell(row=start_row, column=1, value=row['Forest'])
                    ws.cell(row=start_row, column=2, value=row['Truck IDs'])
                    ws.cell(row=start_row, column=3, value=row['Total Trucks'])
                    for col in range(1, 4):
                        cell = ws.cell(row=start_row, column=col)
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                        cell.border = border
                    start_row += 1
                # Box up the day's table
                end_row = start_row - 1
                for r in range(header_row, end_row + 1):
                    for c in range(1, 4):
                        ws.cell(row=r, column=c).border = border
                # Add an empty row between days
                start_row += 1
            # Adjust column widths for grouped daily plan, add extra space
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 4, 40) + 7
            # Row heights for grouped daily plan: set all to 72
            for row_idx in range(1, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 72

        excel_data = buffer.getvalue()
    st.download_button(
        label="Download All Results as Excel (.xlsx)",
        data=excel_data,
        file_name="arise_trucking_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.title(f"{ship_emojis[0]} Truck Forest Allocation Optimizer")
    st.markdown("<span style='font-size:1.1em; color:#111; font-weight:500; background:transparent; display:block;'>Please upload both CSV files and select a season.</span>", unsafe_allow_html=True) 
    